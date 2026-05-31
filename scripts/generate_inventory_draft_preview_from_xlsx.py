#!/usr/bin/env python3
"""从供应商 Excel 生成 VapeStockHub 草稿预览文件。"""

from __future__ import annotations

import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


PROJECT_ROOT = Path("/Users/dezuo/Desktop/项目开发/vapestockhub")
SOURCE_XLSX = PROJECT_ROOT / "exports" / "信息明细0522.xlsx"
OUTPUT_JSON = PROJECT_ROOT / "exports" / "信息明细0522_draft_preview.json"
OUTPUT_MD = PROJECT_ROOT / "exports" / "信息明细0522_draft_preview.md"


# 中文注释：这批表格里只有部分 sheet 提供英文口味，中文口味这里做保守人工映射。
CN_FLAVOR_MAP = {
    "树莓柠檬": "Raspberry Lemon",
    "蓝色拉兹": "Blue Razz",
    "桃子橙子": "Peach Orange",
    "酸西瓜糖": "Sour Watermelon Candy",
    "曼妥思薄荷": "Mint Candy",
    "巫师柠檬": "Wizard Lemon",
    "水晶葡萄球": "Crystal Grape",
    "泡泡可乐": "Bubble Cola",
    "柠檬冻结": "Lemon Freeze",
    "草莓菠萝椰奶": "Strawberry Pineapple Coconut Milk",
    "酸苹果冰": "Sour Apple Ice",
    "蓝色拉兹冰": "Blue Razz Ice",
    "西瓜冰": "Watermelon Ice",
    "草莓冰": "Strawberry Ice",
    "蓝莓西瓜": "Blueberry Watermelon",
    "黑莓樱桃": "Blackberry Cherry",
    "迈阿密薄荷": "Miami Mint",
    "白色软糖": "White Gummy",
    "草莓香蕉": "Strawberry Banana",
    "乔治亚桃子冰": "Georgia Peach Ice",
    "草莓芒果": "Strawberry Mango",
    "法布勒斯": "Fabulous",
    "草莓冰淇淋": "Strawberry Ice Cream",
    "三重甜瓜": "Triple Melon",
    "三重莓果": "Triple Berry",
    "樱桃风暴": "Cherry Storm",
    "小熊软糖": "Gummy Bear",
    "曼妥思薄荷糖": "Mint Candy",
    "西瓜梨火龙果": "Watermelon Pear Dragon Fruit",
    "粉红蓝莓冰": "Pink Blueberry Ice",
    "冰泉": "Ice Spring",
    "樱桃树莓": "Cherry Raspberry",
    "(焦糖)烟草": "Caramel Tobacco",
    "焦糖烟草": "Caramel Tobacco",
    "西瓜梨火龙果": "Watermelon Pear Dragon Fruit",
    "柠檬青柠": "Lemon Lime",
    "三重芒果": "Triple Mango",
    "香蕉冰": "Banana Ice",
    "菠萝冰": "Pineapple Ice",
}


WAREHOUSE_MAP = {
    "保税仓": "Bonded Warehouse, China",
    "东莞": "Dongguan, China",
    "工厂原材料": "Factory Stock, China",
    "总部": "Headquarters, China",
}


PRODUCT_TYPE_DISPOSABLE = "Disposable Vape"
PRODUCT_TYPE_KIT = "Vape Kits"
PRODUCT_TYPE_POD = "Pod"


def slugify(value: str) -> str:
    normalized = re.sub(r"[^a-zA-Z0-9]+", "-", value.strip().lower())
    normalized = re.sub(r"-{2,}", "-", normalized).strip("-")
    return normalized or "draft-item"


def split_lines(value: str | None) -> list[str]:
    if not value:
        return []
    return [line.strip() for line in str(value).splitlines() if str(line).strip()]


def parse_spec_block(spec_text: str | None) -> dict[str, str]:
    text = spec_text or ""
    lines = split_lines(text)
    data: dict[str, str] = {}
    for line in lines:
        clean = line.replace("☑", "").strip()
        if "mg/ml" in clean and "nicotine" not in data:
            data["nicotine"] = clean
        puff_match = re.search(r"Puffs:\s*([0-9* ]+)", clean, re.IGNORECASE)
        if puff_match:
            data["puff"] = puff_match.group(1).replace(" ", "")
        capacity_match = re.search(r"Capacity:\s*([0-9+*.mlML ]+)", clean, re.IGNORECASE)
        if capacity_match:
            data["e_liquid"] = capacity_match.group(1).replace(" ", "")
        flavor_count_match = re.search(r"(\d+)\s+flavors", clean, re.IGNORECASE)
        if flavor_count_match:
            data["flavor_count"] = flavor_count_match.group(1)
        battery_match = re.search(r"(\d+\s*mAh.+)", clean, re.IGNORECASE)
        if battery_match:
            data["battery"] = battery_match.group(1).strip()
    return data


def excel_date_to_text(value: Any) -> str:
    if isinstance(value, (int, float)):
        dt = from_excel(value)
        return dt.strftime("%b %Y")
    return str(value).strip() if value else ""


def translate_flavor(value: str) -> str:
    normalized = value.strip()
    if normalized in CN_FLAVOR_MAP:
        return CN_FLAVOR_MAP[normalized]
    return normalized


def normalize_warehouse(value: str | None) -> str:
    raw = (value or "").strip()
    return WAREHOUSE_MAP.get(raw, raw)


def infer_market_and_featured(version_label: str, preferred_market: str | None = None) -> tuple[str, list[str]]:
    version_text = version_label.lower()
    if preferred_market:
        return preferred_market, []
    if "france" in version_text or "法国" in version_text:
        return "France", ["Western Europe"]
    if "usa" in version_text or "美国" in version_text:
        return "USA", ["USA"]
    return "Global", []


def build_raw_text(entry: dict[str, Any]) -> str:
    fields = entry["normalizedFields"]
    parts = [
        f"Source Sheet: {entry['meta']['sheet']}",
        f"Brand: {fields['brand']}",
        f"Title: {fields['title']}",
        f"Product Type: {fields['product_type']}",
        f"Pricing Mode: {fields['pricing_mode']}",
        f"Quantity: {fields['quantity']}",
        f"MOQ: {fields['moq']}",
        f"Market: {fields['market']}",
        f"Warehouse: {fields['warehouse_location']}",
        f"Nicotine: {fields['nicotine']}",
        f"Puff: {fields['puff']}",
        f"E-liquid: {fields['e_liquid']}",
        f"Production Date: {fields['production_date_text']}",
        f"Flavor Count: {len(fields['flavor_tags'])}",
        "Flavor Breakdown:",
        fields["flavor_breakdown"],
    ]
    if fields["pricing_note"]:
        parts.extend(["Pricing Note:", fields["pricing_note"]])
    if fields["market_access_note"]:
        parts.extend(["Market Access Note:", fields["market_access_note"]])
    return "\n".join(part for part in parts if part)


def make_package(
    *,
    sheet: str,
    title: str,
    brand: str,
    product_type: str,
    pricing_mode: str,
    pricing_note: str,
    price: str,
    quantity: int,
    moq: int,
    market: str,
    featured_markets: list[str],
    market_access_note: str,
    warehouse_location: str,
    nicotine: str,
    puff: str,
    e_liquid: str,
    production_date_text: str,
    flavor_tags: list[str],
    flavor_breakdown: str,
    description_summary: str,
    manifest_notes: str,
    missing_fields: list[str] | None = None,
    risk_flags: list[dict[str, str]] | None = None,
    human_review_focus: list[dict[str, str]] | None = None,
    meta: dict[str, Any] | None = None,
) -> dict[str, Any]:
    package = {
        "version": "v1",
        "rawInput": {
            "sourceType": "excel",
            "supplierName": "信息明细0522",
            "submittedAt": None,
            "sourceLabel": SOURCE_XLSX.name,
            "rawText": "",
        },
        "normalizedFields": {
            "title": title,
            "slug": slugify(title),
            "brand": brand,
            "product_type": product_type,
            "pricing_mode": pricing_mode,
            "pricing_note": pricing_note,
            "price": price,
            "quantity": str(quantity),
            "moq": str(moq),
            "market": market,
            "featured_markets": featured_markets,
            "market_access_note": market_access_note,
            "warehouse_location": warehouse_location,
            "nicotine": nicotine,
            "puff": puff,
            "e_liquid": e_liquid,
            "production_date_text": production_date_text,
            "contact_visibility": "contact_required",
            "images": [],
            "flavor_tags": flavor_tags,
            "flavor_breakdown": flavor_breakdown,
            "description_summary": description_summary,
            "manifest_notes": manifest_notes,
        },
        "missingFields": missing_fields or [],
        "riskFlags": risk_flags or [],
        "humanReviewFocus": human_review_focus or [],
        "meta": meta or {"sheet": sheet},
    }
    package["rawInput"]["rawText"] = build_raw_text(package)
    return package


def add_default_image_risk(package: dict[str, Any]) -> None:
    package["riskFlags"].append(
        {
            "code": "image-missing",
            "severity": "low",
            "message": "Image URL is intentionally left blank for manual completion.",
        }
    )
    package["humanReviewFocus"].append(
        {
            "field": "images",
            "reason": "Fill in the real product image URL before import or publish review.",
        }
    )


def extract_flavor_from_spec(spec_text: str) -> str:
    match = re.search(r"(?:一次性小烟|电子烟烟弹|套装)\s+(.+?)\s+\d+\s*mg", spec_text, re.IGNORECASE)
    return match.group(1).strip() if match else spec_text.strip()


def sum_flavors(rows: list[tuple[str, int]]) -> tuple[list[str], str]:
    flavor_tags: list[str] = []
    breakdown_lines: list[str] = []
    for flavor_name, qty in rows:
        translated = translate_flavor(flavor_name)
        flavor_tags.append(translated)
        breakdown_lines.append(f"{translated}: {qty:,} pcs")
    return flavor_tags, "\n".join(breakdown_lines)


def build_manifest(
    *,
    description_summary: str,
    flavor_breakdown: str,
    packaging_note: str = "",
    pricing_note: str = "",
    market_access_note: str = "",
) -> str:
    sections = [description_summary]
    if pricing_note:
        sections.append(f"Pricing Note:\n{pricing_note}")
    if market_access_note:
        sections.append(f"Market Access Note:\n{market_access_note}")
    if flavor_breakdown:
        sections.append(flavor_breakdown)
    if packaging_note:
        sections.append(f"Packaging / Carton:\n{packaging_note}")
    return "\n\n".join(section.strip() for section in sections if section.strip())


def parse_headquarter_214(ws) -> list[dict[str, Any]]:
    header_note_a = str(ws["A1"].value or "").strip()
    header_note_b = str(ws["A2"].value or "").strip()
    groups: dict[tuple[str, str, str], list[tuple[str, int]]] = defaultdict(list)
    for row in ws.iter_rows(min_row=5, values_only=True):
        if not row or not row[1] or not isinstance(row[3], (int, float)):
            continue
        model_label = str(row[1]).strip()
        spec_text = str(row[2]).strip()
        flavor_cn = extract_flavor_from_spec(spec_text)
        nicotine_match = re.search(r"(\d+\s*mg/ml)", spec_text, re.IGNORECASE)
        nicotine = nicotine_match.group(1).replace(" ", "") if nicotine_match else "50mg/ml"
        version = "Standard Version"
        if "美国版" in spec_text:
            version = "USA Version"
        elif "法国衍生版" in spec_text:
            version = "France Derived Version"
        elif "法国定制版" in spec_text:
            version = "France Custom Version"
        elif "法国版" in spec_text:
            version = "France Version"
        groups[(model_label, version, nicotine)].append((flavor_cn, int(row[3])))

    packages: list[dict[str, Any]] = []
    for (model_label, version, nicotine), flavor_rows in groups.items():
        model_code_match = re.match(r"(UD[0-9A-Z-]+)", model_label)
        model_code = model_code_match.group(1) if model_code_match else model_label
        quantity = sum(qty for _, qty in flavor_rows)
        flavor_tags, flavor_breakdown = sum_flavors(flavor_rows)
        market, featured_markets = infer_market_and_featured(version)
        description_summary = (
            f"Finished stock available from headquarters inventory for {model_code} {version.lower()}. "
            f"Packaging change is not supported for this batch."
        )
        pricing_note = "Current finished stock is USD 1.80 per unit. New production can be quoted separately at USD 5.50 per unit."
        package = make_package(
            sheet=ws.title,
            title=f"{model_code} {nicotine.replace('/ml', '')} Disposable Vape - {version}",
            brand="Pending Brand",
            product_type=PRODUCT_TYPE_DISPOSABLE,
            pricing_mode="exact_price",
            pricing_note=pricing_note,
            price="1.8",
            quantity=quantity,
            moq=1,
            market=market,
            featured_markets=featured_markets,
            market_access_note="Packaging change is not available for the current finished stock batch.",
            warehouse_location="Headquarters, China",
            nicotine=nicotine,
            puff="",
            e_liquid="",
            production_date_text="",
            flavor_tags=flavor_tags,
            flavor_breakdown=flavor_breakdown,
            description_summary=description_summary,
            manifest_notes=build_manifest(
                description_summary=description_summary,
                flavor_breakdown=flavor_breakdown,
                pricing_note=pricing_note,
                market_access_note="Packaging change is not available for the current finished stock batch.",
            ),
            missing_fields=["Brand"],
            risk_flags=[
                {
                    "code": "brand-missing",
                    "severity": "high",
                    "message": "The sheet provides internal UD model codes but does not provide an explicit brand name.",
                },
                {
                    "code": "warehouse-generalized",
                    "severity": "medium",
                    "message": "Warehouse location is generalized as Headquarters, China because the source sheet does not include a city.",
                },
            ],
            human_review_focus=[
                {
                    "field": "brand",
                    "reason": "Confirm the real public-facing brand before converting this preview into a final draft.",
                },
                {
                    "field": "description",
                    "reason": f"Check the original notes: {header_note_a} / {header_note_b}",
                },
            ],
            meta={
                "sheet": ws.title,
                "group": model_label,
                "version": version,
                "source_notes": [header_note_a, header_note_b],
            },
        )
        add_default_image_risk(package)
        packages.append(package)
    return packages


def collect_market_note(rows: list[tuple[str, ...]]) -> str:
    notes: list[str] = []
    for row in rows:
        for cell in row[8:]:
            if isinstance(cell, str):
                clean = cell.strip()
                if clean and clean not in notes:
                    notes.append(clean)
    translated = []
    for note in notes:
        if "必须销出美国" in note:
            translated.append("This version must be sold out of the US and is not allowed for domestic US resale.")
        elif "窜货" in note:
            translated.append("Cross-market diversion to the US is not allowed for this version.")
        elif "保税仓出货要求" in note:
            translated.append("Bonded warehouse shipments require full-carton handling.")
        elif "每个型号至少整箱" in note:
            translated.append("Each SKU should move in full-carton quantity without split cartons.")
        elif "价格暂无更新" in note or "暂无" in note:
            translated.append("Pricing remains case by case for this batch.")
        elif "美国：MD到SD价格8.25" in note:
            translated.append("Reference note in source sheet mentions a US MD-to-SD price of 8.25 for internal review only.")
        elif "包装定制" in note:
            translated.append("Customized packaging version requires at least 2K pcs per flavor.")
        elif "烟油定制" in note:
            translated.append("Flavor-style customization with packaging customization requires at least 5K pcs per flavor.")
    return "\n".join(translated)


def parse_nexa_274(ws) -> list[dict[str, Any]]:
    brand = str(ws["B1"].value or "").strip()
    model = str(ws["E1"].value or "").strip()
    spec_data = parse_spec_block(str(ws["B4"].value or ""))
    packaging_note = str(ws["E4"].value or "").replace("☑", "").strip()
    warehouse = normalize_warehouse(str(ws["B5"].value or ""))
    header_quantity = int(ws["B3"].value or 0)

    france_rows = [tuple(cell for cell in row[:12]) for row in ws.iter_rows(min_row=11, max_row=25, values_only=True)]
    usa_rows = [tuple(cell for cell in row[:12]) for row in ws.iter_rows(min_row=26, max_row=40, values_only=True)]

    def build_version_package(version_name: str, rows: list[tuple[Any, ...]], model_label: str, production_text: str) -> dict[str, Any]:
        flavor_rows: list[tuple[str, int]] = []
        for row in rows:
            if not row or not row[2] or not isinstance(row[4], (int, float)):
                continue
            flavor_rows.append((extract_flavor_from_spec(str(row[2])), int(row[4])))
        quantity = sum(qty for _, qty in flavor_rows)
        flavor_tags, flavor_breakdown = sum_flavors(flavor_rows)
        market, featured = infer_market_and_featured(version_name)
        pricing_note = "Pricing is handled on inquiry because the workbook only contains mixed reference notes, not a clean publishable unit price."
        market_access_note = collect_market_note(rows)
        description_summary = (
            f"{brand} {model} {version_name} finished stock is available in the bonded warehouse. "
            f"Batch production is {production_text} with {quantity:,} pcs across {len(flavor_rows)} flavors."
        )
        package = make_package(
            sheet=ws.title,
            title=f"{brand} {model} 40000 Disposable Vape - {model_label}",
            brand=brand,
            product_type=PRODUCT_TYPE_DISPOSABLE,
            pricing_mode="inquiry_only",
            pricing_note=pricing_note,
            price="0",
            quantity=quantity,
            moq=100,
            market=market,
            featured_markets=featured,
            market_access_note=market_access_note,
            warehouse_location=warehouse,
            nicotine=spec_data.get("nicotine", ""),
            puff=spec_data.get("puff", ""),
            e_liquid=spec_data.get("e_liquid", ""),
            production_date_text=production_text,
            flavor_tags=flavor_tags,
            flavor_breakdown=flavor_breakdown,
            description_summary=description_summary,
            manifest_notes=build_manifest(
                description_summary=description_summary,
                flavor_breakdown=flavor_breakdown,
                packaging_note=packaging_note,
                pricing_note=pricing_note,
                market_access_note=market_access_note,
            ),
            meta={
                "sheet": ws.title,
                "header_quantity": header_quantity,
                "group_quantity": quantity,
                "version": model_label,
            },
        )
        add_default_image_risk(package)
        return package

    return [
        build_version_package("France Version", france_rows, "France Version", "Mar 2026"),
        build_version_package("USA Version", usa_rows, "USA Version", "Oct 2025"),
    ]


def parse_nexa_253(ws) -> list[dict[str, Any]]:
    kit_brand = str(ws["D1"].value or "").strip()
    kit_model = str(ws["G1"].value or "").strip()
    kit_spec = parse_spec_block(str(ws["D4"].value or ""))
    kit_packaging = str(ws["G4"].value or "").replace("☑", "").strip()
    kit_warehouse = normalize_warehouse(str(ws["D5"].value or ""))

    pod_brand = str(ws["D8"].value or "").strip()
    pod_model = str(ws["G8"].value or "").strip()
    pod_spec = parse_spec_block(str(ws["D11"].value or ""))
    pod_packaging = str(ws["G11"].value or "").replace("☑", "").strip()
    pod_warehouse = normalize_warehouse(str(ws["D12"].value or ""))

    kit_new_rows = [tuple(row[:12]) for row in ws.iter_rows(min_row=16, max_row=27, values_only=True)]
    kit_mid_rows = [tuple(row[:12]) for row in ws.iter_rows(min_row=28, max_row=39, values_only=True)]
    pod_new_rows = [tuple(row[:12]) for row in ws.iter_rows(min_row=43, max_row=54, values_only=True)]
    pod_mid_rows = [tuple(row[:12]) for row in ws.iter_rows(min_row=55, max_row=66, values_only=True)]

    def build_group(
        *,
        brand: str,
        model: str,
        rows: list[tuple[Any, ...]],
        qty_index: int,
        spec_data: dict[str, str],
        packaging_note: str,
        warehouse: str,
        version_label: str,
        product_type: str,
        title_suffix: str,
        production_text: str,
        price_hint: str = "",
    ) -> dict[str, Any]:
        flavor_rows: list[tuple[str, int]] = []
        for row in rows:
            if not row or not row[2] or not isinstance(row[qty_index], (int, float)):
                continue
            flavor_rows.append((extract_flavor_from_spec(str(row[2])), int(row[qty_index])))
        quantity = sum(qty for _, qty in flavor_rows)
        flavor_tags, flavor_breakdown = sum_flavors(flavor_rows)
        market, featured = infer_market_and_featured(version_label)
        pricing_note = price_hint or "Pricing is handled on inquiry because the source workbook does not expose one clean unit price for direct publishing."
        market_access_note = collect_market_note(rows)
        description_summary = (
            f"{brand} {model} {title_suffix} stock is available in the bonded warehouse. "
            f"Batch production is {production_text} with {quantity:,} pcs across {len(flavor_rows)} flavors."
        )
        package = make_package(
            sheet=ws.title,
            title=f"{brand} {model} {title_suffix}",
            brand=brand,
            product_type=product_type,
            pricing_mode="inquiry_only",
            pricing_note=pricing_note,
            price="0",
            quantity=quantity,
            moq=100 if product_type == PRODUCT_TYPE_KIT else 200,
            market=market,
            featured_markets=featured,
            market_access_note=market_access_note,
            warehouse_location=warehouse,
            nicotine=spec_data.get("nicotine", ""),
            puff=spec_data.get("puff", ""),
            e_liquid=spec_data.get("e_liquid", ""),
            production_date_text=production_text,
            flavor_tags=flavor_tags,
            flavor_breakdown=flavor_breakdown,
            description_summary=description_summary,
            manifest_notes=build_manifest(
                description_summary=description_summary,
                flavor_breakdown=flavor_breakdown,
                packaging_note=packaging_note,
                pricing_note=pricing_note,
                market_access_note=market_access_note,
            ),
            risk_flags=(
                [
                    {
                        "code": "header-quantity-broken",
                        "severity": "medium",
                        "message": "The top sheet quantity for FLEX KIT is broken as #REF!, so this preview uses the summed flavor rows instead.",
                    }
                ]
                if model == kit_model
                else []
            ),
            human_review_focus=(
                [
                    {
                        "field": "quantity",
                        "reason": "Confirm the summed flavor-row quantity against the supplier source because the sheet header contains a formula error.",
                    }
                ]
                if model == kit_model
                else []
            ),
            meta={
                "sheet": ws.title,
                "version": title_suffix,
                "quantity_from_rows": quantity,
            },
        )
        add_default_image_risk(package)
        return package

    return [
        build_group(
            brand=kit_brand,
            model=kit_model,
            rows=kit_new_rows,
            qty_index=3,
            spec_data=kit_spec,
            packaging_note=kit_packaging,
            warehouse=kit_warehouse,
            version_label="USA New Version",
            product_type=PRODUCT_TYPE_KIT,
            title_suffix="40000 Kit - USA New Version",
            production_text="Late Oct 2025",
        ),
        build_group(
            brand=kit_brand,
            model=kit_model,
            rows=kit_mid_rows,
            qty_index=3,
            spec_data=kit_spec,
            packaging_note=kit_packaging,
            warehouse=kit_warehouse,
            version_label="USA Mid Version",
            product_type=PRODUCT_TYPE_KIT,
            title_suffix="40000 Kit - USA Mid Version",
            production_text="Late Nov 2025",
            price_hint="Source sheet shows internal tier references around USD 4.50-4.65 for the kit, but the pricing structure is incomplete, so this preview keeps the listing as inquiry-only.",
        ),
        build_group(
            brand=pod_brand,
            model=pod_model,
            rows=pod_new_rows,
            qty_index=3,
            spec_data=pod_spec,
            packaging_note=pod_packaging,
            warehouse=pod_warehouse,
            version_label="USA New Version",
            product_type=PRODUCT_TYPE_POD,
            title_suffix="40000 Pod - USA New Version",
            production_text="Late Oct 2025",
        ),
        build_group(
            brand=pod_brand,
            model=pod_model,
            rows=pod_mid_rows,
            qty_index=3,
            spec_data=pod_spec,
            packaging_note=pod_packaging,
            warehouse=pod_warehouse,
            version_label="USA Mid Version",
            product_type=PRODUCT_TYPE_POD,
            title_suffix="40000 Pod - USA Mid Version",
            production_text="Late Nov 2025",
            price_hint="Source sheet shows internal tier references around USD 2.00-2.15 for the pod, but the pricing structure is incomplete, so this preview keeps the listing as inquiry-only.",
        ),
    ]


def parse_english_flavor_sheet(
    ws,
    *,
    brand_cell: str,
    model_cell: str,
    quantity_cell: str,
    production_cell: str,
    spec_cell: str,
    packaging_cell: str,
    warehouse_cell: str,
    product_type: str,
    title_suffix: str,
    market: str = "Global",
    featured_markets: list[str] | None = None,
) -> list[dict[str, Any]]:
    brand = str(ws[brand_cell].value or "").strip()
    model = str(ws[model_cell].value or "").strip()
    quantity = int(ws[quantity_cell].value or 0)
    production = excel_date_to_text(ws[production_cell].value)
    spec_data = parse_spec_block(str(ws[spec_cell].value or ""))
    packaging_note = str(ws[packaging_cell].value or "").replace("☑", "").strip()
    warehouse = normalize_warehouse(str(ws[warehouse_cell].value or ""))

    flavor_rows: list[tuple[str, int]] = []
    for row in ws.iter_rows(min_row=9, values_only=True):
        if len(row) < 7:
            continue
        flavor_en = row[4]
        qty = row[6]
        if isinstance(flavor_en, str) and isinstance(qty, (int, float)):
            flavor_rows.append((str(flavor_en).strip(), int(qty)))

    flavor_tags, flavor_breakdown = sum_flavors(flavor_rows)
    pricing_note = "Pricing is handled on inquiry because the workbook provides stock only and no clean publishable unit price."
    description_summary = (
        f"{brand} {model} stock is available with {quantity:,} pcs across {len(flavor_rows)} flavors. "
        f"Production reference is {production} and the batch is ready for inquiry-based quoting."
    )
    package = make_package(
        sheet=ws.title,
        title=f"{brand} {model} {title_suffix}".strip(),
        brand=brand,
        product_type=product_type,
        pricing_mode="inquiry_only",
        pricing_note=pricing_note,
        price="0",
        quantity=quantity,
        moq=100,
        market=market,
        featured_markets=featured_markets or [],
        market_access_note="",
        warehouse_location=warehouse,
        nicotine=spec_data.get("nicotine", ""),
        puff=spec_data.get("puff", ""),
        e_liquid=spec_data.get("e_liquid", ""),
        production_date_text=production,
        flavor_tags=flavor_tags,
        flavor_breakdown=flavor_breakdown,
        description_summary=description_summary,
        manifest_notes=build_manifest(
            description_summary=description_summary,
            flavor_breakdown=flavor_breakdown,
            packaging_note=packaging_note,
            pricing_note=pricing_note,
        ),
        meta={
            "sheet": ws.title,
            "header_quantity": quantity,
            "flavor_row_count": len(flavor_rows),
        },
    )
    add_default_image_risk(package)
    return [package]


def parse_monster(ws) -> list[dict[str, Any]]:
    packages = parse_english_flavor_sheet(
        ws,
        brand_cell="C2",
        model_cell="F2",
        quantity_cell="C4",
        production_cell="F4",
        spec_cell="C5",
        packaging_cell="F5",
        warehouse_cell="C6",
        product_type=PRODUCT_TYPE_DISPOSABLE,
        title_suffix="Disposable Vape",
        market="Global",
    )
    packages[0]["humanReviewFocus"].append(
        {
            "field": "nicotine",
            "reason": "The sheet lists 20mg/ml and 50mg/ml together, so confirm whether this should become one mixed draft or split by nicotine version later.",
        }
    )
    return packages


def render_markdown(packages: list[dict[str, Any]]) -> str:
    lines = [
        "# 信息明细0522 草稿预览",
        "",
        f"- Source: `{SOURCE_XLSX.name}`",
        f"- Draft Count: `{len(packages)}`",
        "- Note: Image URLs are intentionally left blank for manual completion.",
        "",
    ]
    for index, package in enumerate(packages, start=1):
        fields = package["normalizedFields"]
        lines.extend(
            [
                f"## Draft {index}",
                "",
                f"- Title: `{fields['title']}`",
                f"- Brand: `{fields['brand']}`",
                f"- Product Type: `{fields['product_type']}`",
                f"- Pricing Mode: `{fields['pricing_mode']}`",
                f"- Price USD: `{fields['price']}`",
                f"- Quantity: `{fields['quantity']}`",
                f"- MOQ: `{fields['moq']}`",
                f"- Market: `{fields['market']}`",
                f"- Featured Markets: `{', '.join(fields['featured_markets']) or '-'}`",
                f"- Warehouse: `{fields['warehouse_location']}`",
                f"- Nicotine: `{fields['nicotine'] or '-'}`",
                f"- Puff: `{fields['puff'] or '-'}`",
                f"- E-liquid: `{fields['e_liquid'] or '-'}`",
                f"- Production Date: `{fields['production_date_text'] or '-'}`",
                f"- Flavor Count: `{len(fields['flavor_tags'])}`",
                f"- Missing Fields: `{', '.join(package['missingFields']) or '-'}`",
                f"- Risk Flags: `{ ' | '.join(flag['message'] for flag in package['riskFlags']) or '-' }`",
                "",
                "### Description Summary",
                "",
                fields["description_summary"] or "-",
                "",
                "### Pricing Note",
                "",
                fields["pricing_note"] or "-",
                "",
                "### Market Access Note",
                "",
                fields["market_access_note"] or "-",
                "",
                "### Flavor Tags",
                "",
                ", ".join(fields["flavor_tags"]) or "-",
                "",
                "### Flavor Breakdown",
                "",
                "```text",
                fields["flavor_breakdown"] or "-",
                "```",
                "",
            ]
        )
    return "\n".join(lines).strip() + "\n"


def main() -> None:
    workbook = load_workbook(SOURCE_XLSX, data_only=True)
    packages: list[dict[str, Any]] = []

    packages.extend(parse_headquarter_214(workbook["总部-214口味明细"]))
    packages.extend(parse_nexa_274(workbook["保税仓-274口味明细"]))
    packages.extend(parse_nexa_253(workbook["保税仓-253口味明细"]))
    packages.extend(
        parse_english_flavor_sheet(
            workbook["Crystal Pro Suonon Switch 30K  "],
            brand_cell="C2",
            model_cell="F2",
            quantity_cell="C4",
            production_cell="F4",
            spec_cell="C5",
            packaging_cell="F5",
            warehouse_cell="C6",
            product_type=PRODUCT_TYPE_KIT,
            title_suffix="30000 Kit",
            market="Global",
        )
    )
    packages.extend(
        parse_english_flavor_sheet(
            workbook["CP Pro Suonon Switch 30K POD"],
            brand_cell="C2",
            model_cell="F2",
            quantity_cell="C4",
            production_cell="F4",
            spec_cell="C5",
            packaging_cell="F5",
            warehouse_cell="C6",
            product_type=PRODUCT_TYPE_POD,
            title_suffix="30000 Pod",
            market="Global",
        )
    )
    packages.extend(parse_monster(workbook["Monster 25000"]))
    packages.extend(
        parse_english_flavor_sheet(
            workbook["Dual Galaxy 30000"],
            brand_cell="C2",
            model_cell="F2",
            quantity_cell="C4",
            production_cell="F4",
            spec_cell="C5",
            packaging_cell="F5",
            warehouse_cell="C6",
            product_type=PRODUCT_TYPE_DISPOSABLE,
            title_suffix="Disposable Vape",
            market="Global",
        )
    )
    packages.extend(
        parse_english_flavor_sheet(
            workbook["ARGUSBAR NEON 2K"],
            brand_cell="C2",
            model_cell="F2",
            quantity_cell="C4",
            production_cell="F4",
            spec_cell="C5",
            packaging_cell="F5",
            warehouse_cell="C6",
            product_type=PRODUCT_TYPE_DISPOSABLE,
            title_suffix="Disposable Vape",
            market="Global",
        )
    )

    OUTPUT_JSON.write_text(json.dumps(packages, ensure_ascii=False, indent=2), encoding="utf-8")
    OUTPUT_MD.write_text(render_markdown(packages), encoding="utf-8")

    print(f"generated {len(packages)} draft previews")
    print(OUTPUT_JSON)
    print(OUTPUT_MD)


if __name__ == "__main__":
    main()
