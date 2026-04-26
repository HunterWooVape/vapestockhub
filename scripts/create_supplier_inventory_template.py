from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


# 使用统一的边框样式，方便快速扫读表格。
thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def style_cell(cell, *, bold=False, color="000000", fill=None, align=None, italic=False):
    cell.font = Font(name="Arial", size=10, bold=bold, italic=italic, color=color)
    cell.border = thin_border
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = align


def build_workbook(config, output_path: Path):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    required_fill = PatternFill("solid", fgColor="DDEBF7")
    optional_fill = PatternFill("solid", fgColor="F3F6F9")
    internal_fill = PatternFill("solid", fgColor="E2F0D9")
    guide_fill = PatternFill("solid", fgColor="4F81BD")
    example_fill = PatternFill("solid", fgColor="FFF2CC")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap = Alignment(vertical="top", wrap_text=True)

    wb = Workbook()

    supplier_ws = wb.active
    supplier_ws.title = config["sheet_names"]["supplier"]
    supplier_ws.freeze_panes = "A2"

    for col_idx, (name, level, note) in enumerate(config["supplier_columns"], start=1):
        cell = supplier_ws.cell(row=1, column=col_idx, value=name)
        style_cell(cell, bold=True, color="FFFFFF", fill=header_fill, align=center)

        hint_cell = supplier_ws.cell(row=3, column=col_idx, value=config["hint_formatter"](level, note))
        style_cell(
            hint_cell,
            italic=True,
            color="666666",
            fill=required_fill if level == config["required_label"] else optional_fill,
            align=wrap,
        )

    for col_idx, value in enumerate(config["example_values"], start=1):
        cell = supplier_ws.cell(row=2, column=col_idx, value=value)
        style_cell(cell, fill=example_fill, align=wrap)

    for row in range(4, 104):
        for col_idx in range(1, len(config["supplier_columns"]) + 1):
            style_cell(supplier_ws.cell(row=row, column=col_idx), align=wrap)

    for col, width in config["supplier_widths"].items():
        supplier_ws.column_dimensions[col].width = width
    supplier_ws.row_dimensions[1].height = 24
    supplier_ws.row_dimensions[2].height = 44
    supplier_ws.row_dimensions[3].height = 34

    product_type_dv = DataValidation(
        type="list",
        formula1='"Disposable,Pod System,Kit,E-liquid,Accessory"',
        allow_blank=True,
    )
    image_source_dv = DataValidation(
        type="list",
        formula1='"excel,chat,drive,url,zip"',
        allow_blank=True,
    )
    supplier_ws.add_data_validation(product_type_dv)
    supplier_ws.add_data_validation(image_source_dv)
    product_type_dv.add("F4:F104")
    image_source_dv.add("T4:T104")

    review_ws = wb.create_sheet(config["sheet_names"]["review"])
    review_ws.freeze_panes = "A2"
    for col_idx, (name, _, note) in enumerate(config["review_columns"], start=1):
        header_cell = review_ws.cell(row=1, column=col_idx, value=name)
        style_cell(header_cell, bold=True, color="FFFFFF", fill=header_fill, align=center)

        note_cell = review_ws.cell(row=2, column=col_idx, value=note)
        style_cell(note_cell, italic=True, color="666666", fill=internal_fill, align=wrap)

    for row in range(3, 103):
        for col_idx in range(1, len(config["review_columns"]) + 1):
            style_cell(review_ws.cell(row=row, column=col_idx), align=wrap)

    for col, width in config["review_widths"].items():
        review_ws.column_dimensions[col].width = width
    review_ws.row_dimensions[1].height = 24
    review_ws.row_dimensions[2].height = 34

    grade_dv = DataValidation(type="list", formula1='"Green,Yellow,Red"', allow_blank=True)
    yes_no_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    raw_status_dv = DataValidation(type="list", formula1='"missing,received,extracted"', allow_blank=True)
    review_status_dv = DataValidation(
        type="list",
        formula1='"raw_only,need_optimize,ready_for_upload,uploaded"',
        allow_blank=True,
    )
    review_ws.add_data_validation(grade_dv)
    review_ws.add_data_validation(yes_no_dv)
    review_ws.add_data_validation(raw_status_dv)
    review_ws.add_data_validation(review_status_dv)
    grade_dv.add("E3:E103")
    yes_no_dv.add("I3:I103")
    yes_no_dv.add("L3:L103")
    raw_status_dv.add("J3:J103")
    review_status_dv.add("K3:K103")

    guide_ws = wb.create_sheet(config["sheet_names"]["guide"])
    for col_idx, header in enumerate(config["guide_headers"], start=1):
        cell = guide_ws.cell(row=1, column=col_idx, value=header)
        style_cell(cell, bold=True, color="FFFFFF", fill=guide_fill, align=center)

    for row_idx, row in enumerate(config["guide_rows"], start=2):
        for col_idx, value in enumerate(row, start=1):
            style_cell(guide_ws.cell(row=row_idx, column=col_idx, value=value), align=wrap)

    for col, width in config["guide_widths"].items():
        guide_ws.column_dimensions[col].width = width
    guide_ws.row_dimensions[1].height = 24

    wb.save(output_path)
    print(output_path)


def main():
    root = Path(__file__).resolve().parents[1]
    output_dir = root / "docs" / "templates"
    output_dir.mkdir(parents=True, exist_ok=True)
    supplier_columns_en = [
        ("Supplier Name", "Required", "Supplier company name"),
        ("Contact Name", "Optional", "Primary contact person"),
        ("Contact Channel", "Optional", "WhatsApp / Telegram / WeChat / Email"),
        ("Brand", "Required", "Brand only, no model in this field"),
        ("Model / Product Name", "Required", "Main model name"),
        ("Product Type", "Required", "Disposable / Pod System / Kit / E-liquid / Accessory"),
        ("Unit Price (USD)", "Recommended", "Number only"),
        ("Available Qty", "Required", "Total available quantity"),
        ("MOQ", "Recommended", "Minimum order quantity"),
        ("Target Market", "Required", "Main target region"),
        ("Warehouse Location", "Required", "City, Country/Region"),
        ("Puff Count", "Recommended", "Number only if available"),
        ("Nicotine Strength", "Recommended", "e.g. 5% or 50mg"),
        ("E-liquid Capacity", "Recommended", "e.g. 18ml"),
        ("Flavor List", "Recommended", "Comma-separated flavor overview"),
        ("Flavor Breakdown", "Optional", "Flavor by quantity, can be multi-line"),
        ("Stock Notes", "Recommended", "Raw stock / dispatch / clearance notes"),
        ("Packaging Notes", "Optional", "Box / carton details"),
        ("Extra Notes", "Optional", "Any extra comments"),
        ("Image Source Type", "Recommended", "excel / chat / drive / url / zip"),
        ("Image Reference", "Recommended", "File name, link, folder name, or message reference"),
        ("Image Notes", "Optional", "Photo condition or source note"),
    ]
    review_columns_en = [
        ("Row ID", "Required", "Match row number or internal ID"),
        ("Supplier Name", "Required", "Copy from Supplier_Input"),
        ("Brand", "Required", "Copy from Supplier_Input"),
        ("Model / Product Name", "Required", "Copy from Supplier_Input"),
        ("Intake Grade", "Required", "Green / Yellow / Red"),
        ("Missing Fields", "Optional", "List missing critical fields"),
        ("Risk Flags", "Optional", "Key risk notes"),
        ("Human Review Focus", "Optional", "What to check first"),
        ("Draft Ready", "Required", "Yes / No"),
        ("Image Raw Status", "Required", "missing / received / extracted"),
        ("Image Review Status", "Required", "raw_only / need_optimize / ready_for_upload / uploaded"),
        ("Publish Ready", "Required", "Yes / No"),
        ("Reviewer Notes", "Optional", "Internal notes only"),
    ]
    example_values_en = [
        "Shenzhen ABC Trading",
        "Allen",
        "WhatsApp +971...",
        "Vozol",
        "Star 10000",
        "Disposable",
        "3.20",
        "5000",
        "500",
        "Middle East",
        "Dubai, UAE",
        "10000",
        "5%",
        "18ml",
        "Blue Razz Ice, Watermelon Ice, Mint",
        "Blue Razz Ice - 1200 pcs\nWatermelon Ice - 1500 pcs\nMint - 800 pcs",
        "Ready stock in Dubai warehouse. Mixed flavors available.",
        "10 pcs/box, 200 pcs/carton",
        "Clearance batch, fast-moving stock.",
        "excel",
        "embedded photo from supplier file",
        "Real device photo, large image",
    ]
    supplier_widths = {
        "A": 22,
        "B": 16,
        "C": 24,
        "D": 16,
        "E": 24,
        "F": 18,
        "G": 14,
        "H": 14,
        "I": 12,
        "J": 18,
        "K": 20,
        "L": 12,
        "M": 16,
        "N": 16,
        "O": 28,
        "P": 32,
        "Q": 36,
        "R": 24,
        "S": 24,
        "T": 18,
        "U": 28,
        "V": 26,
    }
    review_widths = {
        "A": 10,
        "B": 22,
        "C": 16,
        "D": 24,
        "E": 14,
        "F": 26,
        "G": 26,
        "H": 28,
        "I": 12,
        "J": 16,
        "K": 24,
        "L": 14,
        "M": 28,
    }
    guide_rows_en = [
        ("Supplier_Input", "Supplier Name", "Required", "Identify supplier source", "Shenzhen ABC Trading", "Use company name."),
        ("Supplier_Input", "Brand", "Required", "Brand identity", "Vozol", "Brand only, do not mix model."),
        ("Supplier_Input", "Model / Product Name", "Required", "Main model", "Star 10000", "Main product name only."),
        ("Supplier_Input", "Product Type", "Required", "Type classification", "Disposable", "Use dropdown values where possible."),
        ("Supplier_Input", "Unit Price (USD)", "Recommended", "Trade unit price", "3.20", "Number only, no currency text."),
        ("Supplier_Input", "Available Qty", "Required", "Available stock quantity", "5000", "Number only when possible."),
        ("Supplier_Input", "MOQ", "Recommended", "Minimum order quantity", "500", "Use number when possible."),
        ("Supplier_Input", "Target Market", "Required", "Main region", "Middle East", "Prefer region wording instead of country list."),
        ("Supplier_Input", "Warehouse Location", "Required", "Stock location", "Dubai, UAE", "Prefer City, Country/Region."),
        ("Supplier_Input", "Flavor List", "Recommended", "Flavor overview", "Blue Razz Ice, Mint", "Comma-separated short list."),
        ("Supplier_Input", "Flavor Breakdown", "Optional", "Flavor by quantity", "Blue Razz Ice - 300 pcs", "One line per flavor if available."),
        ("Supplier_Input", "Stock Notes", "Recommended", "Raw stock context", "Ready stock, sealed carton", "Raw notes are acceptable."),
        ("Supplier_Input", "Image Source Type", "Recommended", "How images arrived", "excel", "Use excel/chat/drive/url/zip."),
        ("Supplier_Input", "Image Reference", "Recommended", "Where to find the image", "embedded photo from supplier file", "Do not rely on embedded images as final media."),
        ("Internal_Review", "Intake Grade", "Required", "Initial intake decision", "Green", "Green = can enter AI draft flow."),
        ("Internal_Review", "Missing Fields", "Optional", "Critical missing fields", "price, warehouse", "Use short machine-readable phrases."),
        ("Internal_Review", "Risk Flags", "Optional", "Material risks", "large image only, real-photo quality", "Keep factual, not emotional."),
        ("Internal_Review", "Human Review Focus", "Optional", "Priority review area", "verify quantity and image relevance", "Guide final manual check."),
        ("Internal_Review", "Draft Ready", "Required", "Ready for AI draft conversion", "Yes", "Only Yes when minimum fields are usable."),
        ("Internal_Review", "Publish Ready", "Required", "Ready to publish", "No", "Only Yes after draft review and image upload."),
    ]
    guide_widths = {"A": 18, "B": 24, "C": 12, "D": 24, "E": 24, "F": 46}

    supplier_columns_cn = [
        ("供应商名称", "必填", "供应商公司名称"),
        ("联系人", "选填", "主要联系人"),
        ("联系方式", "选填", "WhatsApp / Telegram / 微信 / 邮箱"),
        ("品牌", "必填", "只填品牌，不要把型号写在这里"),
        ("型号 / 产品名", "必填", "主要型号名称"),
        ("产品类型", "必填", "Disposable / Pod System / Kit / E-liquid / Accessory"),
        ("单价（USD）", "推荐", "只填数字"),
        ("可售数量", "必填", "当前可售总量"),
        ("MOQ", "推荐", "最小起订量"),
        ("目标市场", "必填", "主要目标区域"),
        ("仓库位置", "必填", "城市, 国家/地区"),
        ("口数", "推荐", "如有则填数字"),
        ("尼古丁浓度", "推荐", "例如 5% 或 50mg"),
        ("烟油容量", "推荐", "例如 18ml"),
        ("口味列表", "推荐", "用逗号分隔的口味概览"),
        ("口味明细", "选填", "每个口味对应数量，可多行"),
        ("库存备注", "推荐", "原始库存 / 发货 / 清仓备注"),
        ("包装备注", "选填", "盒装 / 箱装信息"),
        ("额外备注", "选填", "其他补充说明"),
        ("图片来源类型", "推荐", "excel / chat / drive / url / zip"),
        ("图片引用", "推荐", "文件名、链接、文件夹名或聊天引用"),
        ("图片备注", "选填", "图片状况或来源说明"),
    ]
    review_columns_cn = [
        ("行 ID", "必填", "对应 Excel 行号或内部编号"),
        ("供应商名称", "必填", "从 Supplier_Input 复制"),
        ("品牌", "必填", "从 Supplier_Input 复制"),
        ("型号 / 产品名", "必填", "从 Supplier_Input 复制"),
        ("准入分级", "必填", "Green / Yellow / Red"),
        ("缺失字段", "选填", "列出关键缺失项"),
        ("风险标记", "选填", "关键风险说明"),
        ("人工复核重点", "选填", "优先检查什么"),
        ("可转 Draft", "必填", "Yes / No"),
        ("原始图片状态", "必填", "missing / received / extracted"),
        ("图片处理状态", "必填", "raw_only / need_optimize / ready_for_upload / uploaded"),
        ("可发布", "必填", "Yes / No"),
        ("审核备注", "选填", "仅内部使用"),
    ]
    example_values_cn = [
        "深圳 ABC Trading",
        "Allen",
        "WhatsApp +971...",
        "Vozol",
        "Star 10000",
        "Disposable",
        "3.20",
        "5000",
        "500",
        "Middle East",
        "Dubai, UAE",
        "10000",
        "5%",
        "18ml",
        "Blue Razz Ice, Watermelon Ice, Mint",
        "Blue Razz Ice - 1200 pcs\nWatermelon Ice - 1500 pcs\nMint - 800 pcs",
        "迪拜现货，可混装口味。",
        "10 pcs/box, 200 pcs/carton",
        "清仓批次，走货较快。",
        "excel",
        "供应商表格内嵌图片",
        "实拍图，文件较大",
    ]
    guide_rows_cn = [
        ("Supplier_Input", "供应商名称", "必填", "识别资料来源", "深圳 ABC Trading", "填写公司名称。"),
        ("Supplier_Input", "品牌", "必填", "品牌识别", "Vozol", "只写品牌，不要混入型号。"),
        ("Supplier_Input", "型号 / 产品名", "必填", "主体型号", "Star 10000", "填写主要产品名。"),
        ("Supplier_Input", "产品类型", "必填", "类型归类", "Disposable", "尽量使用下拉值。"),
        ("Supplier_Input", "单价（USD）", "推荐", "交易单价", "3.20", "只填数字，不要带币种文本。"),
        ("Supplier_Input", "可售数量", "必填", "当前库存量", "5000", "尽量使用数字。"),
        ("Supplier_Input", "MOQ", "推荐", "最小起订量", "500", "尽量使用数字。"),
        ("Supplier_Input", "目标市场", "必填", "主目标区域", "Middle East", "优先使用区域口径。"),
        ("Supplier_Input", "仓库位置", "必填", "库存位置", "Dubai, UAE", "优先使用 城市, 国家/地区。"),
        ("Supplier_Input", "口味列表", "推荐", "口味概览", "Blue Razz Ice, Mint", "使用逗号分隔。"),
        ("Supplier_Input", "口味明细", "选填", "口味加数量", "Blue Razz Ice - 300 pcs", "有则每行一个口味。"),
        ("Supplier_Input", "库存备注", "推荐", "原始库存上下文", "Ready stock, sealed carton", "允许原始运营备注。"),
        ("Supplier_Input", "图片来源类型", "推荐", "图片从哪里来", "excel", "使用 excel/chat/drive/url/zip。"),
        ("Supplier_Input", "图片引用", "推荐", "如何找到图片", "供应商表格内嵌图片", "不要把嵌入图当成最终媒体资产。"),
        ("Internal_Review", "准入分级", "必填", "初步准入判断", "Green", "Green 代表可进入 AI draft 流程。"),
        ("Internal_Review", "缺失字段", "选填", "关键缺失项", "price, warehouse", "用简短、可识别短语记录。"),
        ("Internal_Review", "风险标记", "选填", "重要风险", "only real photo, image large", "保持客观事实描述。"),
        ("Internal_Review", "人工复核重点", "选填", "优先复核点", "verify quantity and image relevance", "指向最终人工检查重点。"),
        ("Internal_Review", "可转 Draft", "必填", "是否可进入 AI draft", "Yes", "只有最小字段可用时才填 Yes。"),
        ("Internal_Review", "可发布", "必填", "是否可发布", "No", "补图并完成人工审核后再填 Yes。"),
    ]

    common_config = {
        "supplier_widths": supplier_widths,
        "review_widths": review_widths,
        "guide_widths": guide_widths,
    }

    build_workbook(
        {
            **common_config,
            "sheet_names": {"supplier": "Supplier_Input", "review": "Internal_Review", "guide": "Field_Guide"},
            "supplier_columns": supplier_columns_en,
            "review_columns": review_columns_en,
            "example_values": example_values_en,
            "guide_headers": ["Sheet", "Column", "Level", "Purpose", "Example", "Rule / Note"],
            "guide_rows": guide_rows_en,
            "required_label": "Required",
            "hint_formatter": lambda level, note: f"[{level}] {note}",
        },
        output_dir / "Supplier_Inventory_Template_V1.xlsx",
    )

    build_workbook(
        {
            **common_config,
            "sheet_names": {"supplier": "供应商填写", "review": "内部审核", "guide": "字段说明"},
            "supplier_columns": supplier_columns_cn,
            "review_columns": review_columns_cn,
            "example_values": example_values_cn,
            "guide_headers": ["工作表", "字段", "级别", "用途", "示例", "填写说明"],
            "guide_rows": guide_rows_cn,
            "required_label": "必填",
            "hint_formatter": lambda level, note: f"【{level}】{note}",
        },
        output_dir / "Supplier_Inventory_Template_V1_CN.xlsx",
    )


if __name__ == "__main__":
    main()
